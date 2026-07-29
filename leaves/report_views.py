"""Views για ενότητα Αναφορές (χειριστές αδειών)."""
import io
from datetime import date

from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
from django.shortcuts import render
from django.views.generic import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from accounts.department_utils import KEDASY_KEPEA_DEPARTMENT_TYPE_CODES, SDEY_DEPARTMENT_TYPE_CODES
from accounts.models import Department
from leaves.models import LeaveRequest


def _require_handler(user):
    if not (user.is_leave_handler or user.is_administrator):
        raise PermissionDenied('Μόνο χειριστές αδειών έχουν πρόσβαση.')


class ReportsIndexView(LoginRequiredMixin, TemplateView):
    """Κεντρική σελίδα Αναφορών."""
    template_name = 'leaves/reports_index.html'

    def dispatch(self, request, *args, **kwargs):
        _require_handler(request.user)
        return super().dispatch(request, *args, **kwargs)


@login_required
def department_managers_report(request):
    """Αναφορά προϊσταμένων ανά τμήμα — επισημαίνει τμήματα χωρίς προϊστάμενο."""
    _require_handler(request.user)

    departments = (
        Department.objects.select_related(
            'manager', 'department_type', 'parent_department', 'prefecture',
        )
        .order_by('name')
    )
    rows = []
    missing_count = 0
    for dept in departments:
        manager = dept.manager
        if not manager:
            missing_count += 1
        rows.append({
            'department': dept,
            'manager': manager,
            'has_manager': bool(manager),
        })

    return render(request, 'leaves/report_department_managers.html', {
        'rows': rows,
        'total_count': len(rows),
        'missing_count': missing_count,
        'with_manager_count': len(rows) - missing_count,
    })


# Τύποι τμημάτων που ΕΞΑΙΡΟΥΝΤΑΙ από την εξαγωγή για SCH
_SCH_EXCLUDED_DEPT_TYPES = KEDASY_KEPEA_DEPARTMENT_TYPE_CODES + SDEY_DEPARTMENT_TYPE_CODES


def _sch_export_queryset(date_from: date, date_to: date):
    """
    Επιστρέφει LeaveRequest που πληρούν τα κριτήρια εξαγωγής SCH:
    - Ολοκληρωμένες αιτήσεις (COMPLETED)
    - Υποβλήθηκαν (created_at ή η πρώτη περίοδος) εντός date_from..date_to
    - Ο αιτών ΔΕΝ ανήκει σε ΚΕΔΑΣΥ/ΚΕΠΕΑ/ΣΔΕΥ
    """
    excluded_depts = Department.objects.filter(
        department_type__code__in=_SCH_EXCLUDED_DEPT_TYPES
    ).values_list('id', flat=True)

    qs = (
        LeaveRequest.objects
        .filter(status='COMPLETED')
        .exclude(user__department__id__in=excluded_depts)
        .select_related(
            'user__department__department_type',
            'leave_type',
        )
        .prefetch_related('periods')
    )

    # Φιλτράρισμα βάσει ημερομηνίας έναρξης πρώτης περιόδου
    qs = qs.filter(
        periods__start_date__gte=date_from,
        periods__start_date__lte=date_to,
    ).distinct()

    return qs.order_by('user__last_name', 'user__first_name', 'id')


@login_required
def sch_export_report(request):
    """Αναφορά εξαγωγής για SCH (αριθμός μητρώου, τύπος άδειας, πρωτόκολλο)."""
    _require_handler(request.user)

    date_from_str = request.GET.get('date_from', '')
    date_to_str = request.GET.get('date_to', '')
    export = request.GET.get('export', '')

    date_from = None
    date_to = None
    rows = []
    errors = []

    if date_from_str and date_to_str:
        try:
            date_from = date.fromisoformat(date_from_str)
            date_to = date.fromisoformat(date_to_str)
            if date_from > date_to:
                errors.append('Η ημερομηνία έναρξης πρέπει να είναι πριν από τη λήξη.')
            else:
                qs = _sch_export_queryset(date_from, date_to)
                for lr in qs:
                    first_period = lr.periods.order_by('start_date').first()
                    rows.append({
                        'employee_number': lr.user.employee_number or '',
                        'full_name': lr.user.full_name,
                        'leave_type': lr.leave_type.name,
                        'start_date': first_period.start_date if first_period else '',
                        'total_days': lr.total_days,
                        'pdede_protocol_number': lr.pdede_protocol_number or '',
                        'pdede_protocol_date': lr.pdede_protocol_date.date() if lr.pdede_protocol_date else '',
                    })
        except ValueError:
            errors.append('Μη έγκυρη μορφή ημερομηνίας.')
    elif date_from_str or date_to_str:
        errors.append('Συμπληρώστε και τις δύο ημερομηνίες.')

    if export == 'xlsx' and rows and not errors:
        return _sch_export_xlsx(rows, date_from, date_to)

    return render(request, 'leaves/report_sch_export.html', {
        'rows': rows,
        'errors': errors,
        'date_from': date_from_str,
        'date_to': date_to_str,
        'total_count': len(rows),
    })


def _sch_export_xlsx(rows, date_from, date_to):
    """Δημιουργεί και επιστρέφει αρχείο Excel για την εξαγωγή SCH."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Εξαγωγή SCH'

    # Στυλ
    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    title_font = Font(bold=True, size=13)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Τίτλος
    ws.merge_cells('A1:G1')
    ws['A1'] = (
        f'Εξαγωγή για SCH  |  Περίοδος: {date_from.strftime("%d/%m/%Y")} – {date_to.strftime("%d/%m/%Y")}'
    )
    ws['A1'].font = title_font
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 24

    # Κεφαλίδες
    headers = [
        'Αρ. Μητρώου SCH',
        'Ονοματεπώνυμο',
        'Τύπος Άδειας',
        'Ημ/νία Έναρξης',
        'Εγκριθείσες Ημέρες',
        'Αρ. Πρωτοκόλλου Εισερχομένων',
        'Ημ/νία Πρωτοκόλλου Εισερχομένων',
    ]
    col_widths = [20, 30, 30, 18, 18, 32, 32]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 30

    # Δεδομένα
    for row_idx, row in enumerate(rows, start=3):
        start_date = row['start_date']
        proto_date = row['pdede_protocol_date']

        ws.cell(row=row_idx, column=1, value=row['employee_number'])
        ws.cell(row=row_idx, column=2, value=row['full_name'])
        ws.cell(row=row_idx, column=3, value=row['leave_type'])
        ws.cell(row=row_idx, column=4, value=start_date.strftime('%d/%m/%Y') if start_date else '')
        ws.cell(row=row_idx, column=5, value=row['total_days'])
        ws.cell(row=row_idx, column=6, value=row['pdede_protocol_number'])
        ws.cell(row=row_idx, column=7, value=proto_date.strftime('%d/%m/%Y') if proto_date else '')

        # Εναλλακτικό χρώμα γραμμών
        if row_idx % 2 == 0:
            fill = PatternFill('solid', fgColor='DCE6F1')
            for col_idx in range(1, 8):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    # Freeze πάνω από τα δεδομένα
    ws.freeze_panes = 'A3'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f'sch_export_{date_from}_{date_to}.xlsx'
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response
